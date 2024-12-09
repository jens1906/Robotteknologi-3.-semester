import os
import sys
import subprocess

import cv2
import numpy as np

#import main
import xlsxwriter
from openpyxl import load_workbook

project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
sys.path.append(project_root)
from Palette_detection import LocateChecker as lc


# Load the image
#imageBefore = cv2.imread('Objective_testing\Objective_testing_before.jpg')
#imageAfter = cv2.imread('Objective_testing\Objective_testing_after.jpg')


def OPSNR(imgX, imgY):

    return round(cv2.PSNR(imgX, imgY), 2)

def MeanBrightnessError(imgX, imgY):
    #In the start the before and after picture get converted into grey scale picture
    #This is done because of greyscale also can be seen at as brightness levels
    imgXGrey = cv2.cvtColor(imgX, cv2.COLOR_BGR2GRAY)
    imgYGrey = cv2.cvtColor(imgY, cv2.COLOR_BGR2GRAY)

    #After the convertion the formula can be calculated for the 2 pictures by doing as seen down under.
    result = round(np.mean(imgXGrey)-np.mean(imgYGrey), 2)
    return result 


def AverageGradient(imgX):
    #Firstly the images har converted into greyscale so that there is less information to work with.
    #This can be done because it is wanted to see how big changes there is in the picture, aka sharpness
    #These differences can be found in brightness levels and therefore greyscale is great
    imgGrey = cv2.cvtColor(imgX, cv2.COLOR_BGR2GRAY)

    #The sobel command finds all of the gradients, this works in the format
    #cv2.Sobel(ImageInput, ResolutionOfData, xdirection, ydirection, kernelsize)
    #This means that it here will make a array where it has [x gradient, y gradient]
    gradients = [cv2.Sobel(imgGrey, cv2.CV_64F, 1, 0, ksize=3), cv2.Sobel(imgGrey, cv2.CV_64F, 0, 1, ksize=3)]

    #Now the magnitude of the gradients are found which is done by finding the length of the gradient vectors.
    gradientMagnitutes = np.sqrt(gradients[0]**2 + gradients[1]**2)

    #Then the average gradient can be found
    avgGradientResult = np.sum(gradientMagnitutes)/((imgGrey.shape[1]-1)*(imgGrey.shape[0]-1))
    
    avgGradientResult = np.round(avgGradientResult, decimals=2)
    return avgGradientResult

def ReadyExcel(worksheet):
    # Set the values in the cells using openpyxl's .cell() method
    worksheet.cell(row=1, column=1, value='Filename')

    # PSNR
    worksheet.cell(row=1, column=2, value='PSNR Ground checker diff Reference checker')
    worksheet.cell(row=1, column=3, value='PSNR Ground checker diff Enhanced checker')

    # MBE
    worksheet.cell(row=1, column=4, value='MBE Ground diff Reference')
    worksheet.cell(row=1, column=5, value='MBE Ground diff Enhanced')
    worksheet.cell(row=1, column=6, value='MBE Ground diff Dehazed')

    # AG
    worksheet.cell(row=1, column=7, value='AG Ground')
    worksheet.cell(row=1, column=8, value='AG Reference')
    worksheet.cell(row=1, column=9, value='AG Enhanced')
    worksheet.cell(row=1, column=10, value='AG Dehazed')

    return

def OTDatacollection(folder, Foldername = None):
    if '/' in folder:
        Parentfolder = folder.rsplit('/', 1)[0]
        if Foldername == None:
            FolderName = folder.rsplit('/', 1)[-1]
    elif '\\' in folder:
        Parentfolder = folder.rsplit('\\', 1)[0] 
        if Foldername == None:   
            FolderName = folder.rsplit('\\', 1)[-1]            
    #print("Parentfolder:", Parentfolder)
    if Foldername != None:
        if '/' in folder:
            Parentfolder = folder
            FolderName = Foldername.rsplit('/', 1)[-1]
        elif '\\' in folder:
            FolderName = Foldername.rsplit('\\', 1)[-1]   
            Parentfolder = folder   
    
    worksheet = None
    ExcelFile = f'{Parentfolder}/AllOTResults.xlsx'
    if not os.path.isfile(ExcelFile):
        workbook = xlsxwriter.Workbook(ExcelFile)
        worksheet = workbook.add_worksheet(FolderName)
        workbook.close()
        workbook = load_workbook(ExcelFile)
        worksheet = workbook.active
    else:
        workbook = load_workbook(ExcelFile)
        if FolderName not in workbook.sheetnames:
            worksheet = workbook.create_sheet(FolderName)
        else:
            SheetLoc = workbook[FolderName]
            workbook.remove(SheetLoc)
            worksheet = workbook.create_sheet(FolderName)
    ReadyExcel(worksheet)
    return workbook, worksheet, ExcelFile

def ObjectiveTesting(File, Improved, reference, worksheet, dehazed, enhancedChecker, referenceChecker):
    reference = cv2.imread(reference)
    Original = cv2.imread("P3\Results\Data\GroundTruth\Beside_Camera_AutoTarget5_light5_exp29311.0_20242211_103548.png")
    OriginalChecker = cv2.resize( cv2.imread("P3\Palette_detection\Colour_checker_from_Vikki_full.png")[170: 997, 520: 1705], (enhancedChecker.shape[1],enhancedChecker.shape[0]), interpolation=cv2.INTER_AREA)
    #OriginalChecker = cv2.resize( cv2.imread("P3\Palette_detection\Colour_checker_from_Vikki_full_test_enviorment.png")[39: 211, 108: 360], (enhancedChecker.shape[1],enhancedChecker.shape[0]), interpolation=cv2.INTER_AREA)
    referenceChecker = cv2.cvtColor(referenceChecker, cv2.COLOR_BGR2RGB)
    enhancedChecker = cv2.cvtColor(enhancedChecker, cv2.COLOR_BGR2RGB)

    try:
        PsnrGroundVSReference = OPSNR(OriginalChecker, referenceChecker)
        #PsnrGroundVSReference = OPSNR(OriginalChecker, cv2.filter2D(referenceChecker, -1, np.ones((5,5),np.float32)/25))          
        #PsnrGroundVSReference = OPSNR(cv2.cvtColor(OriginalChecker, cv2.COLOR_BGR2GRAY), cv2.cvtColor(referenceChecker, cv2.COLOR_BGR2GRAY))

        PsnrGroundVSEnhanced = OPSNR(OriginalChecker, enhancedChecker)
        #PsnrGroundVSEnhanced = OPSNR(OriginalChecker, cv2.filter2D(enhancedChecker, -1, np.ones((5,5),np.float32)/25))
        #PsnrGroundVSEnhanced = OPSNR(cv2.cvtColor(OriginalChecker, cv2.COLOR_BGR2GRAY), cv2.cvtColor(enhancedChecker, cv2.COLOR_BGR2GRAY))
        
    except:
        print("PSNR FAIL")
        print(OriginalChecker.shape)
        print(referenceChecker.shape)
        print(enhancedChecker.shape)
        pass
    #print("PSNR Ground vs Reference:", PsnrGroundVSReference)
    #print("PSNR Ground vs Enhanced:", PsnrGroundVSEnhanced)
    #cv2.imshow('Original Image', cv2.resize(OriginalChecker, (0,0), fx=1, fy=1))
    #cv2.imshow('reference Image', cv2.resize(referenceChecker, (0,0), fx=1, fy=1))
    #cv2.imshow('dehazed Checker', cv2.resize(enhancedChecker, (0,0), fx=1, fy=1))
    #cv2.waitKey(0)
    MBEGroundVSReference = MeanBrightnessError(Original, reference)
    MBEGroundVSEnhanced = MeanBrightnessError(Original, Improved)
    MBEGroundVsDehazed = MeanBrightnessError(Original, dehazed)

    AGGround = AverageGradient(Original)
    AGReference = AverageGradient(reference)
    AGEnhanced = AverageGradient(Improved)
    AGDehazed = AverageGradient(dehazed)

    next_row = worksheet.max_row + 1    
    #print(next_row)
    #
    #PSNR
    #
    worksheet.cell(row=next_row, column=1, value=File)                      #Filename
    worksheet.cell(row=next_row, column=2, value=PsnrGroundVSReference)     #PSNR Ground checker vs Reference checker
    worksheet.cell(row=next_row, column=3, value=PsnrGroundVSEnhanced)      #PSNR Ground checker vs Enhanced checker

    #
    #MBE
    #
    worksheet.cell(row=next_row, column=4, value=MBEGroundVSReference)      #MBE Ground vs Reference
    worksheet.cell(row=next_row, column=5, value=MBEGroundVSEnhanced)       #MBE Ground vs Enhanced
    worksheet.cell(row=next_row, column=6, value=MBEGroundVsDehazed)        #MBE Ground vs Dehazed

    #
    #AG
    #
    worksheet.cell(row=next_row, column=7, value=AGGround)                  #AG Ground
    worksheet.cell(row=next_row, column=8, value=AGReference)               #AG Reference
    worksheet.cell(row=next_row, column=9, value=AGEnhanced)                #AG Enhanced
    worksheet.cell(row=next_row, column=10, value=AGDehazed)                #AG Dehazed
    return 

def ObjectiveTestingFail(File, worksheet):
    next_row = worksheet.max_row + 1    
    #print(next_row)
    worksheet.cell(row=next_row, column=1, value=File)  # Column A

    return 

def average(worksheet):
    next_row = worksheet.max_row + 1
    worksheet.cell(row=next_row, column=1, value='=COUNT(B2:B' + str(next_row - 1) + ') & " / " & '+ str(next_row - 2) + '')
    worksheet.cell(row=next_row, column=2, value="=AVERAGE(B2:B" + str(next_row-1) + ")")
    worksheet.cell(row=next_row, column=3, value="=AVERAGE(C2:C" + str(next_row-1) + ")")
    worksheet.cell(row=next_row, column=4, value="=AVERAGE(D2:D" + str(next_row-1) + ")")
    worksheet.cell(row=next_row, column=5, value="=AVERAGE(E2:E" + str(next_row-1) + ")")
    worksheet.cell(row=next_row, column=6, value="=AVERAGE(F2:F" + str(next_row-1) + ")")
    worksheet.cell(row=next_row, column=7, value="=AVERAGE(G2:G" + str(next_row-1) + ")")
    worksheet.cell(row=next_row, column=8, value="=AVERAGE(H2:H" + str(next_row-1) + ")")
    worksheet.cell(row=next_row, column=9, value="=AVERAGE(I2:I" + str(next_row-1) + ")")
    worksheet.cell(row=next_row, column=10, value="=AVERAGE(J2:J" + str(next_row-1) + ")")
    

def AdjustExcel(worksheet):
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter  # Get the column letter (e.g., A, B, C)
        for cell in column_cells:
            try:
                if cell.value:  # Check if the cell has a value
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2  # Add some padding to the width
        worksheet.column_dimensions[column_letter].width = adjusted_width
    return    

